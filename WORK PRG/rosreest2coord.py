import openpyxl
import pandas as pd
from rosreestr2coord import Area
import time
import requests

# Путь к файлу
file_path = '/Users/vetvy/Desktop/L I S T # 4 - B R E S.xlsx'

# Загружаем файл Excel и указываем конкретный лист
wb = openpyxl.load_workbook(file_path)
ws = wb["- L I S T -"]

# Читаем файл с помощью pandas для удобства работы с данным листом
df = pd.read_excel(file_path, sheet_name="- L I S T -")

# Создаем столбец "Coord RosRSTR" в DataFrame, если он не существует
if 'Coord RosRSTR' not in df.columns:
    df['Coord RosRSTR'] = ""

# Фильтрация корректных кадастровых номеров
df['Кадастрик'] = df['Кадастрик'].apply(lambda x: str(x).strip() if pd.notna(x) and isinstance(x, str) else None)

# Применяем координаты по кадастровому номеру
total_entries = len(df)
print(f"Всего кадастровых номеров для обработки: {total_entries}")

for index, row in df.iterrows():
    cadastral_number = row['Кадастрик']  # Столбец 'AJ' с кадастровыми номерами

    # Проверка на корректность кадастрового номера
    if not cadastral_number or not cadastral_number.count(':') == 3:
        print(f"Обработан кадастровый номер {index + 1}/{total_entries}: Пропущен (некорректный номер)")
        continue

    attempt = 0
    success = False
    while attempt < 3 and not success:
        try:
            attempt += 1
            print(f"Attempt {attempt}: Start downloading area info for {cadastral_number}")
            
            # Получение координат по кадастровому номеру с использованием параметра -C
            area = Area(cadastral_number, with_log=False, center_only=True)  # center_only=True соответствует параметру -C
            area_coords = area.get_coord()
            
            if isinstance(area_coords, dict) and 'center' in area_coords:
                coords_str = f"{area_coords['center'][0]}, {area_coords['center'][1]}"
                df.at[index, 'Coord RosRSTR'] = coords_str
                ws[f'AM{index + 2}'] = coords_str
                success = True
                print(f"Обработан кадастровый номер {index + 1}/{total_entries}: Координаты {coords_str}")
            else:
                raise ValueError("Unexpected response format")
            
        except requests.exceptions.Timeout: 
            print(f"Timeout error on attempt {attempt} for cadastral number {cadastral_number}")
            time.sleep(1)  # Задержка перед повторной попыткой
            
        except Exception as e:
            if attempt == 3:  # После трех неудачных попыток, зафиксировать ошибку
                error_message = f"Ошибка: {e}"
                df.at[index, 'Coord RosRSTR'] = error_message
                ws[f'AM{index + 2}'] = error_message
                print(f"Обработан кадастровый номер {index + 1}/{total_entries}: {error_message}")
            else:
                print(f"Ошибка на попытке {attempt} для кадастрового номера {cadastral_number}: {e}")
            time.sleep(1)  # Задержка перед повторной попыткой

# Сохранение изменений в файл
wb.save(file_path)
print("Все кадастровые номера успешно обработаны и сохранены в файле.")
