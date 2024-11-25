import openpyxl
import pandas as pd
import requests
import time

# Путь к файлу
file_path = '/Users/vetvy/Desktop/L I S T # 4 - B R E S.xlsx'

# Загружаем файл Excel и указываем конкретный лист
wb = openpyxl.load_workbook(file_path)
ws = wb["- L I S T -"]

# Читаем файл с помощью pandas для удобства работы с данным листом
df = pd.read_excel(file_path, sheet_name="- L I S T -")

# Ваш API ключ DaData
api_key = "a5b1346cb2ca102dc55fdfa7602d195b2a59e217"

# Функция для нормализации адреса через DaData API
def normalize_address(address):
    url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/address"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Token {api_key}"
    }
    data = {"query": address}
    
    try:
        response = requests.post(url, json=data, headers=headers)
        response.raise_for_status()
        
        suggestions = response.json().get('suggestions', [])
        if suggestions:
            return suggestions[0]['unrestricted_value'], None
        else:
            return address, "Невозможно нормализовать"
    except requests.exceptions.RequestException as e:
        return address, f"Ошибка: {e}"

# Функция для получения координат через DaData API
def get_coordinates(address):
    url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/address"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Token {api_key}"
    }
    data = {"query": address}
    
    try:
        response = requests.post(url, json=data, headers=headers)
        response.raise_for_status()
        
        suggestions = response.json().get('suggestions', [])
        if suggestions:
            data = suggestions[0]['data']
            if 'geo_lat' in data and 'geo_lon' in data:
                return (data['geo_lon'], data['geo_lat']), None
        return (None, None), "Координаты не найдены"
    except requests.exceptions.RequestException as e:
        return (None, None), f"Ошибка: {e}"

# Применяем нормализацию и получаем координаты
total_addresses = len(df)
print(f"Всего адресов для обработки: {total_addresses}")

for index, row in df.iterrows():
    # Исходный адрес из столбца "Адрес в Пирамиде"
    original_address = row['Адрес в Пирамиде']
    
    # Нормализуем адрес
    normalized_address, norm_status = normalize_address(original_address)
    df.at[index, 'AK'] = normalized_address
    
    # Получаем координаты
    (lon, lat), coord_status = get_coordinates(normalized_address)
    df.at[index, 'AL'] = f"{lon}, {lat}" if lon and lat else ""
    
    # Выводим статус обработки
    print(f"Обработан адрес {index + 1}/{total_addresses}.")
    
    if norm_status:
        print(f"  Нормализация: {norm_status}")
    if coord_status:
        print(f"  Координаты: {coord_status}")
    
    # Записываем результат в файл
    ws[f'AK{index + 2}'] = normalized_address
    ws[f'AL{index + 2}'] = f"{lon}, {lat}" if lon and lat else ""

    # Пауза, чтобы не перегружать API
    time.sleep(0.5)

# Сохранение изменений в файл
wb.save(file_path)
print("Все адреса успешно обработаны и сохранены в файле.")
